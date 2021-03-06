from io import BytesIO
from zipfile import BadZipFile

from flask import Blueprint, request, jsonify
from openpyxl import load_workbook

from sheet.auth import jwt_required

bp = Blueprint("excel", __name__, url_prefix="/excel")


@bp.route('/info', methods=['POST'])
@jwt_required
def info():
    """Gets the tabs of an Excel file and return alphabetically ordered."""
    if not request.data:
        return jsonify(message='No binary file was sent.'), 400

    try:
        file = BytesIO(request.data)

        wb = load_workbook(file)
        sheet_names = wb.sheetnames
        sheet_names.sort()
    except (KeyError, BadZipFile):
        return jsonify(message='The file is not a Excel file.'), 400

    return jsonify(tabs=sheet_names)
